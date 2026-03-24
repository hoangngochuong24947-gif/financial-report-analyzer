"""
====================================================================
模块名称：test_export.py
模块功能：Excel 导出和图表生成测试
====================================================================
"""

import pytest
from decimal import Decimal
from datetime import date
from io import BytesIO

from openpyxl import load_workbook

from src.models.financial_statements import BalanceSheet, IncomeStatement, CashFlowStatement
from src.models.financial_metrics import (
    ProfitabilityMetrics, SolvencyMetrics, EfficiencyMetrics, DuPontResult
)
from src.export.excel_exporter import ExcelExporter
from src.export.chart_generator import (
    generate_trend_chart,
    generate_ratio_bar_chart,
    generate_dupont_chart,
)


# ============================================================
# 测试数据
# ============================================================

@pytest.fixture
def sample_balance_sheet():
    """测试用资产负债表"""
    return BalanceSheet(
        stock_code="600519",
        report_date=date(2023, 12, 31),
        total_current_assets=Decimal("180000000000.00"),
        total_non_current_assets=Decimal("40000000000.00"),
        total_assets=Decimal("220000000000.00"),
        total_current_liabilities=Decimal("30000000000.00"),
        total_non_current_liabilities=Decimal("5000000000.00"),
        total_liabilities=Decimal("35000000000.00"),
        total_equity=Decimal("185000000000.00"),
    )


@pytest.fixture
def sample_income_statement():
    """测试用利润表"""
    return IncomeStatement(
        stock_code="600519",
        report_date=date(2023, 12, 31),
        total_revenue=Decimal("130000000000.00"),
        operating_cost=Decimal("15000000000.00"),
        operating_profit=Decimal("90000000000.00"),
        total_profit=Decimal("88000000000.00"),
        net_income=Decimal("74000000000.00"),
    )


@pytest.fixture
def sample_cashflow():
    """测试用现金流量表"""
    return CashFlowStatement(
        stock_code="600519",
        report_date=date(2023, 12, 31),
        operating_cashflow=Decimal("60000000000.00"),
        investing_cashflow=Decimal("-10000000000.00"),
        financing_cashflow=Decimal("-20000000000.00"),
        net_cashflow=Decimal("30000000000.00"),
    )


@pytest.fixture
def sample_profitability():
    """测试用盈利能力指标"""
    return ProfitabilityMetrics(
        roe=Decimal("0.4000"),
        roa=Decimal("0.3364"),
        net_profit_margin=Decimal("0.5692"),
        gross_profit_margin=Decimal("0.8846"),
    )


@pytest.fixture
def sample_dupont():
    """测试用杜邦分析结果"""
    return DuPontResult(
        roe=Decimal("0.4000"),
        net_profit_margin=Decimal("0.5692"),
        asset_turnover=Decimal("0.5909"),
        equity_multiplier=Decimal("1.1892"),
    )


# ============================================================
# Excel 导出测试
# ============================================================

class TestExcelExporter:
    """Excel 导出器测试"""

    def test_export_creates_valid_xlsx(
        self, sample_balance_sheet, sample_income_statement, sample_cashflow
    ):
        """测试生成的 Excel 文件是否合法"""
        exporter = ExcelExporter()
        data = exporter.export_full_report(
            stock_code="600519",
            stock_name="贵州茅台",
            balance_sheets=[sample_balance_sheet],
            income_statements=[sample_income_statement],
            cashflow_statements=[sample_cashflow],
        )

        # 能成功加载为 openpyxl Workbook
        wb = load_workbook(BytesIO(data))
        assert len(wb.sheetnames) >= 3  # 至少有三大报表

    def test_balance_sheet_sheet(self, sample_balance_sheet, sample_income_statement, sample_cashflow):
        """测试资产负债表 Sheet 内容"""
        exporter = ExcelExporter()
        data = exporter.export_full_report(
            stock_code="600519",
            stock_name="贵州茅台",
            balance_sheets=[sample_balance_sheet],
            income_statements=[sample_income_statement],
            cashflow_statements=[sample_cashflow],
        )

        wb = load_workbook(BytesIO(data))
        assert "资产负债表" in wb.sheetnames

        ws = wb["资产负债表"]
        # 标题行
        assert "贵州茅台" in str(ws["A1"].value)
        # 数据行存在
        assert ws.cell(row=4, column=1).value is not None

    def test_with_ratios_and_dupont(
        self, sample_balance_sheet, sample_income_statement,
        sample_cashflow, sample_profitability, sample_dupont
    ):
        """测试包含财务比率和杜邦分析的完整报表"""
        exporter = ExcelExporter()
        data = exporter.export_full_report(
            stock_code="600519",
            stock_name="贵州茅台",
            balance_sheets=[sample_balance_sheet],
            income_statements=[sample_income_statement],
            cashflow_statements=[sample_cashflow],
            profitability=sample_profitability,
            dupont=sample_dupont,
        )

        wb = load_workbook(BytesIO(data))
        assert "财务比率" in wb.sheetnames
        assert "杜邦分析" in wb.sheetnames
        assert len(wb.sheetnames) == 5  # 三大报表 + 比率 + 杜邦

    def test_export_returns_bytes(self, sample_balance_sheet, sample_income_statement, sample_cashflow):
        """测试返回值是 bytes 类型"""
        exporter = ExcelExporter()
        data = exporter.export_full_report(
            stock_code="600519",
            stock_name="贵州茅台",
            balance_sheets=[sample_balance_sheet],
            income_statements=[sample_income_statement],
            cashflow_statements=[sample_cashflow],
        )

        assert isinstance(data, bytes)
        assert len(data) > 0


# ============================================================
# 图表生成测试
# ============================================================

class TestChartGenerator:
    """图表生成测试"""

    def test_trend_chart_generates_png(self):
        """测试趋势图生成"""
        png = generate_trend_chart(
            metric_name="净利润（亿元）",
            dates=["2020", "2021", "2022", "2023"],
            values=[466, 524, 592, 640],
        )

        assert isinstance(png, bytes)
        assert len(png) > 1000  # PNG 应该有一定大小
        # 检查 PNG 文件头
        assert png[:4] == b'\x89PNG'

    def test_ratio_bar_chart(self):
        """测试比率柱状图"""
        png = generate_ratio_bar_chart(
            ratios={"ROE": 0.25, "ROA": 0.18, "净利率": 0.30},
            title="盈利能力分析",
        )

        assert isinstance(png, bytes)
        assert png[:4] == b'\x89PNG'

    def test_dupont_chart(self, sample_dupont):
        """测试杜邦分析图"""
        png = generate_dupont_chart(sample_dupont, "贵州茅台")

        assert isinstance(png, bytes)
        assert png[:4] == b'\x89PNG'

    def test_trend_chart_single_point(self):
        """测试单点趋势图（边界情况）"""
        png = generate_trend_chart(
            metric_name="营收",
            dates=["2023"],
            values=[1000],
        )
        assert isinstance(png, bytes)


"""
====================================================================
【运行方式】

cd financial-report-analyzer
poetry run pytest tests/test_export.py -v

====================================================================
"""
