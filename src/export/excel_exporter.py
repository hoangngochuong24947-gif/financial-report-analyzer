"""
====================================================================
模块名称：excel_exporter.py
模块功能：使用 openpyxl 生成专业 Excel 财务报表

【函数接口总览】
┌─────────────────────────────┬────────────────────────────┬─────────────────────────┐
│ 类/函数                      │ 方法                        │ 说明                     │
├─────────────────────────────┼────────────────────────────┼─────────────────────────┤
│ ExcelExporter                │ export_full_report()        │ 导出完整多Sheet报表      │
│                              │ _write_balance_sheet()      │ 写入资产负债表           │
│                              │ _write_income_statement()   │ 写入利润表               │
│                              │ _write_cashflow()           │ 写入现金流量表           │
│                              │ _write_ratios()             │ 写入财务比率             │
│                              │ _write_dupont()             │ 写入杜邦分析             │
└─────────────────────────────┴────────────────────────────┴─────────────────────────┘

【数据流向】
→ 被 api/routes_report.py 调用
→ 输出 .xlsx 文件
====================================================================
"""

import io
from decimal import Decimal
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter

from src.models.financial_statements import BalanceSheet, IncomeStatement, CashFlowStatement
from src.models.financial_metrics import (
    ProfitabilityMetrics, SolvencyMetrics, EfficiencyMetrics, DuPontResult
)
from src.utils.logger import logger


# 样式常量
HEADER_FONT = Font(name="微软雅黑", size=12, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
TITLE_FONT = Font(name="微软雅黑", size=16, bold=True, color="2F5496")
DATA_FONT = Font(name="微软雅黑", size=10)
AMOUNT_FORMAT = '#,##0.00'
RATIO_FORMAT = '0.00%'
THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')


class ExcelExporter:
    """
    Excel 报表导出器

    生成包含多个 Sheet 的专业 Excel 文件：
    - 资产负债表
    - 利润表
    - 现金流量表
    - 财务比率
    - 杜邦分析
    """

    def export_full_report(
        self,
        stock_code: str,
        stock_name: str,
        balance_sheets: List[BalanceSheet],
        income_statements: List[IncomeStatement],
        cashflow_statements: List[CashFlowStatement],
        profitability: Optional[ProfitabilityMetrics] = None,
        solvency: Optional[SolvencyMetrics] = None,
        efficiency: Optional[EfficiencyMetrics] = None,
        dupont: Optional[DuPontResult] = None,
    ) -> bytes:
        """
        导出完整的多 Sheet Excel 报表

        Args:
            stock_code: 股票代码
            stock_name: 股票名称
            balance_sheets: 资产负债表列表（多期）
            income_statements: 利润表列表
            cashflow_statements: 现金流量表列表
            profitability: 盈利能力指标（可选）
            solvency: 偿债能力指标（可选）
            efficiency: 运营效率指标（可选）
            dupont: 杜邦分析结果（可选）

        Returns:
            bytes: Excel 文件二进制内容

        Examples:
            >>> exporter = ExcelExporter()
            >>> data = exporter.export_full_report("600519", "贵州茅台", sheets, stmts, cfs)
            >>> with open("report.xlsx", "wb") as f:
            ...     f.write(data)
        """
        wb = Workbook()

        logger.info(f"Exporting Excel report for {stock_code} ({stock_name})...")

        # 写入各个 Sheet
        self._write_balance_sheet(wb, balance_sheets, stock_name)
        self._write_income_statement(wb, income_statements, stock_name)
        self._write_cashflow(wb, cashflow_statements, stock_name)

        if profitability or solvency or efficiency:
            self._write_ratios(wb, profitability, solvency, efficiency, stock_name)

        if dupont:
            self._write_dupont(wb, dupont, stock_name)

        # 删除默认的空 Sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        # 导出为二进制
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        logger.info(f"Excel report exported: {len(wb.sheetnames)} sheets")
        return buffer.getvalue()

    def _write_balance_sheet(
        self, wb: Workbook, sheets: List[BalanceSheet], stock_name: str
    ):
        """写入资产负债表 Sheet"""
        ws = wb.create_sheet(title="资产负债表")

        # 标题
        ws.merge_cells("A1:H1")
        title_cell = ws["A1"]
        title_cell.value = f"{stock_name} — 资产负债表"
        title_cell.font = TITLE_FONT
        title_cell.alignment = CENTER_ALIGN

        # 表头
        headers = ["报告期", "流动资产合计", "非流动资产合计", "资产总计",
                    "流动负债合计", "非流动负债合计", "负债合计", "所有者权益合计"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER

        # 数据行
        for row_idx, sheet in enumerate(sheets, 4):
            data = [
                str(sheet.report_date),
                float(sheet.total_current_assets),
                float(sheet.total_non_current_assets),
                float(sheet.total_assets),
                float(sheet.total_current_liabilities),
                float(sheet.total_non_current_liabilities),
                float(sheet.total_liabilities),
                float(sheet.total_equity),
            ]
            for col_idx, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = DATA_FONT
                cell.border = THIN_BORDER
                if col_idx > 1:
                    cell.number_format = AMOUNT_FORMAT

        # 自适应列宽
        self._auto_column_width(ws)

    def _write_income_statement(
        self, wb: Workbook, statements: List[IncomeStatement], stock_name: str
    ):
        """写入利润表 Sheet"""
        ws = wb.create_sheet(title="利润表")

        ws.merge_cells("A1:F1")
        title_cell = ws["A1"]
        title_cell.value = f"{stock_name} — 利润表"
        title_cell.font = TITLE_FONT
        title_cell.alignment = CENTER_ALIGN

        headers = ["报告期", "营业总收入", "营业总成本", "营业利润", "利润总额", "净利润"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER

        for row_idx, stmt in enumerate(statements, 4):
            data = [
                str(stmt.report_date),
                float(stmt.total_revenue),
                float(stmt.operating_cost),
                float(stmt.operating_profit),
                float(stmt.total_profit),
                float(stmt.net_income),
            ]
            for col_idx, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = DATA_FONT
                cell.border = THIN_BORDER
                if col_idx > 1:
                    cell.number_format = AMOUNT_FORMAT

        self._auto_column_width(ws)

    def _write_cashflow(
        self, wb: Workbook, statements: List[CashFlowStatement], stock_name: str
    ):
        """写入现金流量表 Sheet"""
        ws = wb.create_sheet(title="现金流量表")

        ws.merge_cells("A1:E1")
        title_cell = ws["A1"]
        title_cell.value = f"{stock_name} — 现金流量表"
        title_cell.font = TITLE_FONT
        title_cell.alignment = CENTER_ALIGN

        headers = ["报告期", "经营活动现金流", "投资活动现金流", "筹资活动现金流", "现金净增加额"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER

        for row_idx, stmt in enumerate(statements, 4):
            data = [
                str(stmt.report_date),
                float(stmt.operating_cashflow),
                float(stmt.investing_cashflow),
                float(stmt.financing_cashflow),
                float(stmt.net_cashflow),
            ]
            for col_idx, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = DATA_FONT
                cell.border = THIN_BORDER
                if col_idx > 1:
                    cell.number_format = AMOUNT_FORMAT

        self._auto_column_width(ws)

    def _write_ratios(
        self,
        wb: Workbook,
        profitability: Optional[ProfitabilityMetrics],
        solvency: Optional[SolvencyMetrics],
        efficiency: Optional[EfficiencyMetrics],
        stock_name: str,
    ):
        """写入财务比率 Sheet"""
        ws = wb.create_sheet(title="财务比率")

        ws.merge_cells("A1:C1")
        title_cell = ws["A1"]
        title_cell.value = f"{stock_name} — 财务比率分析"
        title_cell.font = TITLE_FONT
        title_cell.alignment = CENTER_ALIGN

        headers = ["指标类别", "指标名称", "指标值"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER

        row = 4
        ratio_data = []

        if profitability:
            ratio_data.extend([
                ("盈利能力", "净资产收益率 (ROE)", profitability.roe),
                ("盈利能力", "总资产收益率 (ROA)", profitability.roa),
                ("盈利能力", "销售净利率", profitability.net_profit_margin),
                ("盈利能力", "毛利率", profitability.gross_profit_margin),
            ])

        if solvency:
            ratio_data.extend([
                ("偿债能力", "流动比率", solvency.current_ratio),
                ("偿债能力", "速动比率", solvency.quick_ratio),
                ("偿债能力", "资产负债率", solvency.debt_to_asset_ratio),
                ("偿债能力", "产权比率", solvency.debt_to_equity_ratio),
            ])

        if efficiency:
            ratio_data.extend([
                ("运营效率", "总资产周转率", efficiency.asset_turnover),
                ("运营效率", "权益乘数", efficiency.equity_turnover),
            ])

        for category, name, value in ratio_data:
            ws.cell(row=row, column=1, value=category).font = DATA_FONT
            ws.cell(row=row, column=2, value=name).font = DATA_FONT
            value_cell = ws.cell(row=row, column=3, value=float(value))
            value_cell.font = DATA_FONT
            value_cell.number_format = RATIO_FORMAT
            for col in range(1, 4):
                ws.cell(row=row, column=col).border = THIN_BORDER
            row += 1

        self._auto_column_width(ws)

    def _write_dupont(self, wb: Workbook, dupont: DuPontResult, stock_name: str):
        """写入杜邦分析 Sheet"""
        ws = wb.create_sheet(title="杜邦分析")

        ws.merge_cells("A1:C1")
        title_cell = ws["A1"]
        title_cell.value = f"{stock_name} — 杜邦分析 (ROE 拆解)"
        title_cell.font = TITLE_FONT
        title_cell.alignment = CENTER_ALIGN

        # 公式说明
        formula_cell = ws.cell(row=3, column=1, value="ROE = 销售净利率 × 总资产周转率 × 权益乘数")
        formula_cell.font = Font(name="微软雅黑", size=11, italic=True, color="666666")

        # 表头
        headers = ["拆解因子", "数值"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER

        # 数据
        data = [
            ("净资产收益率 (ROE)", dupont.roe),
            ("销售净利率", dupont.net_profit_margin),
            ("总资产周转率", dupont.asset_turnover),
            ("权益乘数", dupont.equity_multiplier),
        ]

        for row_idx, (name, value) in enumerate(data, 6):
            ws.cell(row=row_idx, column=1, value=name).font = DATA_FONT
            value_cell = ws.cell(row=row_idx, column=2, value=float(value))
            value_cell.font = DATA_FONT
            value_cell.number_format = RATIO_FORMAT
            for col in range(1, 3):
                ws.cell(row=row_idx, column=col).border = THIN_BORDER

        # 验证行
        verify_row = len(data) + 7
        ws.cell(row=verify_row, column=1, value="验证").font = Font(
            name="微软雅黑", size=10, bold=True
        )
        verify_value = float(dupont.net_profit_margin * dupont.asset_turnover * dupont.equity_multiplier)
        ws.cell(row=verify_row, column=2, value=verify_value).number_format = RATIO_FORMAT

        self._auto_column_width(ws)

    @staticmethod
    def _auto_column_width(ws, min_width: int = 12, max_width: int = 30):
        """
        自动调整列宽

        Args:
            ws: 工作表对象
            min_width: 最小列宽
            max_width: 最大列宽
        """
        for col_idx in range(1, ws.max_column + 1):
            max_len = min_width
            col_letter = get_column_letter(col_idx)

            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        cell_len = len(str(cell.value))
                        if cell_len > max_len:
                            max_len = min(cell_len + 2, max_width)

            ws.column_dimensions[col_letter].width = max_len


"""
====================================================================
【使用示例】

from src.export.excel_exporter import ExcelExporter

exporter = ExcelExporter()

# 导出完整报表
data = exporter.export_full_report(
    stock_code="600519",
    stock_name="贵州茅台",
    balance_sheets=balance_sheets,
    income_statements=income_statements,
    cashflow_statements=cashflow_statements,
    profitability=profitability_metrics,
    solvency=solvency_metrics,
    dupont=dupont_result,
)

# 保存文件
with open("茅台_财务分析.xlsx", "wb") as f:
    f.write(data)

====================================================================
"""
